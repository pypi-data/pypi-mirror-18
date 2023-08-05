#!python

import argparse
import json
import logging
import openpyxl
import pathlib
from collections import OrderedDict


def get_args():
    parser = argparse.ArgumentParser(
        description="Convert Chronophore data from json to xlsx."
    )

    parser.add_argument('input', help="path of json file to use as input")

    parser.add_argument(
        '-o', '--output',
        help="path of excel file to generate (default: ./$(input_file_name).xlsx)"
    )
    parser.add_argument(
        '-c', '--clobber', action="store_true",
        help="overwrite output file if it exists"
    )
    parser.add_argument(
        '-v', '--verbose', action="store_true",
        help="print a detailed log"
    )

    return parser.parse_args()


def data_to_excel(data, output_file):
    """Saves data from one of Chronophore's json files
    to an Excel spreadsheet.

    Parameters:
        - data: A dictionary (preferably an OrderedDict) from
                a json file created by Chronophore.
        - title: The title of the output file.
    """
    try:
        headers = {header for key in data for header in data[key].keys()}
        logging.debug("Headers: {}".format(headers))
    except (TypeError, AttributeError) as e:
        logging.error(
            "Invalid data. Must be a valid dictionary: {} ({})".format(data, e)
        )
    else:
        # assign columns to headers
        header_columns = {
            header: column for key in data
            for column, header in enumerate(headers, start=2)
        }
        logging.debug('Header columns: {}'.format(header_columns))

        # create sheet
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = output_file.name

        # make headers
        sheet.cell(row=1, column=1).value = "Key"
        for (header, col_num) in header_columns.items():
            sheet.cell(row=1, column=col_num).value = header

        sheet.freeze_panes = 'A2'

        # fill in data
        for row_num, key in enumerate(data.keys(), start=2):
            sheet.cell(row=row_num, column=1).value = key

            for header, value in data[key].items():
                col_num = header_columns[header]
                sheet.cell(row=row_num, column=col_num).value = value
                logging.debug(
                    'Key:{}, Header:{}, Row:{}, Column:{}, Value:{}'.format(
                        key, header, row_num, col_num, value
                    )
                )

        return wb


if __name__ == '__main__':
    args = get_args()

    if args.verbose:
        LOGGING_LEVEL = logging.DEBUG
    else:
        LOGGING_LEVEL = logging.WARNING

    logging.basicConfig(
        level=LOGGING_LEVEL,
        format='%(levelname)s:%(asctime)s:%(message)s'
    )

    CLOBBER = args.clobber

    JSON_FILE = pathlib.Path(args.input)

    if args.output:
        EXCEL_FILE = pathlib.Path(args.output)
    else:
        EXCEL_FILE = pathlib.Path('.', JSON_FILE.stem + '.xlsx')

    if not CLOBBER and EXCEL_FILE.exists():
        logging.warning(
            "'{}' exists. To overwrite it, use '--clobber'.".format(EXCEL_FILE)
        )
        raise SystemExit

    with JSON_FILE.open('r') as f:
        data = json.load(f, object_pairs_hook=OrderedDict)
    logging.info('Json loaded: {}.'.format(JSON_FILE))

    wb = data_to_excel(data, EXCEL_FILE)
    wb.save(str(EXCEL_FILE))
    if CLOBBER:
        logging.info("Worksheet saved (file overwritten): {}".format(EXCEL_FILE))
    else:
        logging.info("Worksheet saved: {}".format(EXCEL_FILE))
