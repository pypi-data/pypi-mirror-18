from AccessControl import ClassSecurityInfo
from DateTime import DateTime
from Products.ATContentTypes.content import schemata
from Products.Archetypes import atapi
from Products.Archetypes.utils import contentDispositionHeader
from Products.CMFCore.utils import getToolByName
from Products.PloneFormGen.config import DOWNLOAD_SAVED_PERMISSION
from Products.PloneFormGen.content.saveDataAdapter import FormSaveDataAdapter
from Products.PloneFormGen.interfaces.actionAdapter import IPloneFormGenActionAdapter
from ftw.xlsxsaveadapter import _
from ftw.xlsxsaveadapter import PROJECTNAME
from email import Encoders
from email.Header import Header
from email.MIMEBase import MIMEBase
from email.MIMEMultipart import MIMEMultipart
from email.MIMEText import MIMEText
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.writer.excel import save_virtual_workbook
from zope.i18n import translate
from zope.interface import implements


XlsxDataAdapterSchema = FormSaveDataAdapter.schema.copy() + atapi.Schema((
    atapi.StringField(
        name='xlsx_recipients',
        required=False,
        widget=atapi.LinesWidget(
            label=_(u'label_xlsx_recipients', default=u'XLSX recipients'),
            description=_(u'help_xlsx_recipients',
                          default=u'One email address per line'),
        )
    )
))

schemata.finalizeATCTSchema(XlsxDataAdapterSchema, folderish=True,
                            moveDiscussion=False)
schemata.finalizeATCTSchema(XlsxDataAdapterSchema, folderish=True,
                            moveDiscussion=False)
XlsxDataAdapterSchema['DownloadFormat'].widget.visible = {'edit': 0, 'view': 0}

if 'excludeFromNav' in XlsxDataAdapterSchema.keys():
    XlsxDataAdapterSchema['excludeFromNav'].default = True


class XlsxDataAdapter(FormSaveDataAdapter):
    implements(IPloneFormGenActionAdapter)

    meta_type = "XlsxDataAdapterSchema"
    portal_type = "XlsxDataAdapterSchema"
    archetype_name = 'Xlsx Data Adapter'
    schema = XlsxDataAdapterSchema
    security = ClassSecurityInfo()

    immediate_view = 'fg_savedata_view_p3'
    default_view = 'fg_savedata_view_p3'
    suppl_views = ('fg_savedata_tabview_p3', 'fg_savedata_recview_p3',)

    def onSuccess(self, fields, REQUEST=None, loopstop=False):
        """ saves input data and initiates mail"""
        super(XlsxDataAdapter, self).onSuccess(fields, REQUEST, loopstop)
        self.sendXLSX()

    security.declareProtected(DOWNLOAD_SAVED_PERMISSION, 'download_csv')
    def download_csv(self, REQUEST=None, RESPONSE=None, plain=False):
        """ Called by default view. Redirect to xlsx download.
        """
        return self.download_xlsx(REQUEST, RESPONSE, plain)

    security.declareProtected(DOWNLOAD_SAVED_PERMISSION, 'download_xlsx')
    def download_xlsx(self, REQUEST=None, RESPONSE=None, plain=False):
        if not plain:
            filename = '%s.xlsx' % self.id
            header_value = contentDispositionHeader('attachment', 'l1',
                                                    filename=filename)

            RESPONSE.setHeader("Content-Disposition", header_value)
            RESPONSE.setHeader("Content-Type", 'application/vnd.ms-excel')

        display_header = getattr(self, 'UseColumnNames', False)

        values = self.getSavedFormInput()

        if display_header:
            return self.create_export_excel(values, self.getShowColumnTitles())
        return self.create_export_excel(values)

    def getShowColumnTitles(self):
        titles = []
        form = self.aq_parent
        for field in self.getShowFields():
            titles.append(form[field].Title())
        titles.extend(self.getColumnTitles()[-len(self.getExtraData()):])
        return titles

    def sendXLSX(self):
        """ Sends the XLSX File in an email to the xlsx_recipient.
        """

        recipients = self.getXlsx_recipients()

        if not recipients or len(recipients) == 0:
            return

        portal = getToolByName(self, 'portal_url').getPortalObject()
        sender = portal.getProperty('email_from_address', 'noreply@nohost.com')
        event = translate(_(u'entry_added', default=u'Entry was added'),
                          context=self.REQUEST)

        subject = "%s: %s" % (self.Title().decode('utf-8'), event)
        text = translate(_(u'xlsxmail_body',
                         default=u'As attachment you get the new XLSX'),
                         context=self.REQUEST)

        mhost = getToolByName(self, 'MailHost')

        msg_root = MIMEMultipart()
        msg_root['Subject'] = Header(subject, 'iso-8859-1')
        msg_root['From'] = sender
        msg_root['To'] = ', '.join(recipients)
        text_part = MIMEText(text.encode('cp1252'), 'plain', 'cp1252')
        msg_root.attach(text_part)

        xlsx_data = self.download_xlsx(plain=True)

        part = MIMEBase('application', "application/vnd.ms-excel")
        part.set_payload(xlsx_data)
        Encoders.encode_base64(part)
        part.add_header(
            'Content-Disposition', 'attachment; filename="%s_%s.%s"'
            % (self.id, DateTime().Date().replace('/', ''), "xlsx"))
        msg_root.attach(part)

        mhost.send(str(msg_root))

    def create_export_excel(self, data, header=False):
        """ Creates excel and returns it as a string.
            data is a 2d array holding the content.
            header is a array with the column titles indexed identical to data.
        """
        book = Workbook()

        sheet = book.active
        sheet.title = translate(_(u'xlsx_export_sheet_name', default="Form Data"),
                                context=self.REQUEST)

        if header:
            bold = Font(bold=True)
            for column_nr, heading_title in enumerate(header):
                cell = sheet.cell(column=column_nr+1, row=1)
                cell.font = bold
                cell.value = heading_title

        content_offset = 2 if header else 1

        for row_nr, dataentry in enumerate(data):
            for column_nr, dataentry_field in enumerate(dataentry):
                cell = sheet.cell(column=column_nr+1,
                                  row=row_nr+content_offset,
                                  value=dataentry_field)

        # get the maximum amount of characters in a column to estimate the width
        for column in sheet.columns:
            maxwidth = 0
            for cell in column:
                if not cell.value:
                    continue
                cwidth = len(cell.value)
                if cwidth > maxwidth:
                    maxwidth = cwidth
            # a bit more space for readability
            sheet.column_dimensions[cell.column].width = maxwidth + 5

        return save_virtual_workbook(book)


atapi.registerType(XlsxDataAdapter, PROJECTNAME)
