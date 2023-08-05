from openpyxl import load_workbook
import traceback

from ..client import get_client


class Unitcleaner:
    def __init__(self, **data):
        self.data = data


class CleanerConfigurator:
    def __init__(self, cleaner_uid):
        """
        Parameters
        ----------
        cleaner_uid: cleaner_id or (project_id, cleaner_name)
        client: optional
        """

        self.client = get_client()

        if isinstance(cleaner_uid, str):
            self.cleaner_id = cleaner_uid
        else:
            project_id, cleaner_name = cleaner_uid
            cleaners_l = self.client.list(
                "opmeasures_cleaners/cleaners/",
                params=dict(project=project_id, name=cleaner_name))["data"]
            assert len(cleaners_l) == 1, "One and only one cleaner should exist, found %s. " \
                                         "(project id: %s, cleaner_name: %s)" % (len(cleaners_l),
                                                                                 project_id, cleaner_name)
            self.cleaner_id = cleaners_l[0]["id"]

    def configure_unit_cleaner(self, data, update_if_exists=False):
        unitcleaner_l = self.client.list(
            "opmeasures_cleaners/unitcleaners/",
            params=dict(cleaner=self.cleaner_id, external_name=data["external_name"])
        )["data"]
        if len(unitcleaner_l) == 1:
            if not update_if_exists:
                raise AssertionError("unitcleaner already exists, can't create (use update_if_exists=True)")
            # we delete
            self.client.destroy("opmeasures_cleaners/unitcleaners/", unitcleaner_l[0]["id"])

        # we create
        self.client.create("opmeasures_cleaners/unitcleaners/", data)

    def batch_configure(self, xlsx_path, update_if_exists=False):
        unitcleaners = excel_to_platform(xlsx_path)
        for uc in unitcleaners:
            try:
                data = dict(cleaner=self.cleaner_id)
                for k, v in uc.data.items():
                    if v is None:
                        continue
                    data[k] = v
                self.configure_unit_cleaner(data, update_if_exists=update_if_exists)
            except Exception:
                name = getattr(uc, "external_name", "Unknown (no external name provided).")
                print("Error while configuring unitcleaner: %s\n%s\n\n" % (name, traceback.format_exc()))


def excel_to_platform(path, max_input_length=20000):
    unitcleaners = []
    wb = load_workbook(path)
    ws = wb["Series"]
    start_row = 8
    for row in range(start_row, max_input_length + start_row):
        # print(ws['C{}'.format(row)].value)
        if ws['C{}'.format(row)].value == 'x':
            if ws['D{}'.format(row)].value == 'yes':
                ws['D{}'.format(row)].value = 'true'
            elif ws['D{}'.format(row)].value == 'no':
                ws['D{}'.format(row)].value = 'false'
            unitcleaners.append(Unitcleaner(
                external_name=ws['A{}'.format(row)].value,
                name=ws['B{}'.format(row)].value,
                active=ws['D{}'.format(row)].value,
                freq=ws['E{}'.format(row)].value,
                input_unit_type=ws['F{}'.format(row)].value,
                input_convention=ws['G{}'.format(row)].value,
                clock=ws['H{}'.format(row)].value,
                timezone=ws['I{}'.format(row)].value,
                unit=ws['J{}'.format(row)].value,
                label=ws['K{}'.format(row)].value,
                unit_type=ws['L{}'.format(row)].value,
                resample_rule=ws['M{}'.format(row)].value,
                interpolate_limit=ws['N{}'.format(row)].value,
                wait_offset=ws['O{}'.format(row)].value,
                operation_fct=ws['P{}'.format(row)].value,
                filter_fct=ws['Q{}'.format(row)].value,
                derivative_filter_fct=ws['R{}'.format(row)].value,
                custom_delay=ws['S{}'.format(row)].value,
                custom_fct=ws['T{}'.format(row)].value,
                custom_before_offset=ws['U{}'.format(row)].value,
                custom_after_offset=ws['V{}'.format(row)].value
            ))
    return unitcleaners
