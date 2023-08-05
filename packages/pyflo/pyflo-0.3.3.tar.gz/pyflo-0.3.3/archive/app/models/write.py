"""Functions for writing database information to excel spreadsheets.

:copyright: 2016, See AUTHORS for more details
:license: GNU General Public License, See LICENSE for more details.

"""

import warnings

from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment


def storm_tabulation(analysis, results, filename):
    """Create and writes HGL elevation results to an excel spreadsheet.

    Args:
        analysis (pyflo.models.Analysis): The analysis in which data will be extracted.
        results (list[pyflo.models.Result]): The results in which data will be extracted.
        filename(str): The name of the file to create.

    """
    warnings.simplefilter('ignore')
    wb = load_workbook('app/template/stormsewer_template.xlsx')
    warnings.simplefilter('default')

    ws = wb.get_sheet_by_name('TEMPLATE')
    ws.title = analysis.display_name
    ws['Z8'] = analysis.display_name
    ws['Z9'] = analysis.node.display_name
    ws['Z10'] = analysis.last_run.strftime('%a %b %d %Y %H:%M:%S')
    ws['Z11'] = analysis.intensity_polynomial.display_name

    row_start = 12
    row_increment = 3

    for i, result in enumerate(results):
        row_curr = row_start + row_increment * i

        ws.merge_cells('A{0}:A{1}'.format(row_curr + 0, row_curr + 1))
        ws.merge_cells('B{0}:B{1}'.format(row_curr + 0, row_curr + 2))
        ws.merge_cells('C{0}:C{1}'.format(row_curr + 1, row_curr + 2))
        ws.merge_cells('D{0}:D{1}'.format(row_curr + 1, row_curr + 2))
        ws.merge_cells('E{0}:E{1}'.format(row_curr + 1, row_curr + 2))
        ws.merge_cells('F{0}:F{1}'.format(row_curr + 0, row_curr + 2))
        ws.merge_cells('G{0}:G{1}'.format(row_curr + 0, row_curr + 2))
        ws.merge_cells('H{0}:H{1}'.format(row_curr + 0, row_curr + 2))
        ws.merge_cells('I{0}:I{1}'.format(row_curr + 0, row_curr + 2))
        ws.merge_cells('J{0}:J{1}'.format(row_curr + 0, row_curr + 2))
        ws.merge_cells('K{0}:K{1}'.format(row_curr + 0, row_curr + 2))
        ws.merge_cells('L{0}:L{1}'.format(row_curr + 0, row_curr + 2))
        ws.merge_cells('M{0}:M{1}'.format(row_curr + 0, row_curr + 2))
        ws.merge_cells('N{0}:N{1}'.format(row_curr + 0, row_curr + 2))
        ws.merge_cells('R{0}:R{1}'.format(row_curr + 0, row_curr + 2))
        ws.merge_cells('S{0}:S{1}'.format(row_curr + 0, row_curr + 2))
        ws.merge_cells('T{0}:T{1}'.format(row_curr + 0, row_curr + 2))
        ws.merge_cells('U{0}:U{1}'.format(row_curr + 0, row_curr + 1))
        ws.merge_cells('V{0}:V{1}'.format(row_curr + 0, row_curr + 1))
        ws.merge_cells('W{0}:W{1}'.format(row_curr + 0, row_curr + 2))
        ws.merge_cells('X{0}:AA{1}'.format(row_curr + 0, row_curr + 2))

        border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )
        set_cell_and_format(ws, 'A{0}'.format(row_curr + 0), border, result.pipe.node_1.name)
        set_cell_and_format(ws, 'A{0}'.format(row_curr + 2), border, result.pipe.node_2.name)
        set_cell_and_format(ws, 'B{0}'.format(row_curr + 0), border, result.pipe.length)
        set_cell_and_format(ws, 'C{0}'.format(row_curr + 0), border, result.basin.c)
        set_cell_and_format(ws, 'C{0}'.format(row_curr + 1), border, '=E{0}/D{0}'.format(row_curr + 1))
        set_cell_and_format(ws, 'D{0}'.format(row_curr + 0), border, result.basin.area)
        set_cell_and_format(ws, 'D{0}'.format(row_curr + 1), border, result.area)
        set_cell_and_format(ws, 'E{0}'.format(row_curr + 0), border, '=C{0}*D{0}'.format(row_curr + 0))
        set_cell_and_format(ws, 'E{0}'.format(row_curr + 1), border, result.runoff_area)
        set_cell_and_format(ws, 'F{0}'.format(row_curr + 0), border, result.basin.tc)
        set_cell_and_format(ws, 'G{0}'.format(row_curr + 0), border, result.time)
        set_cell_and_format(ws, 'H{0}'.format(row_curr + 0), border, '=B{0}/V{0}/60'.format(row_curr + 0))
        set_cell_and_format(ws, 'I{0}'.format(row_curr + 0), border, result.get_y)
        set_cell_and_format(ws, 'J{0}'.format(row_curr + 0), border, 0.0)
        set_cell_and_format(ws, 'K{0}'.format(row_curr + 0), border, result.flow)
        set_cell_and_format(ws, 'L{0}'.format(row_curr + 0), border, 0.0)
        set_cell_and_format(ws, 'M{0}'.format(row_curr + 0), border, result.pipe.node_1.elevation)
        set_cell_and_format(ws, 'N{0}'.format(row_curr + 0), border, '=M{0}-O{0}'.format(row_curr + 0))
        set_cell_and_format(ws, 'O{0}'.format(row_curr + 0), border, result.hgl_1)
        set_cell_and_format(ws, 'O{0}'.format(row_curr + 1), border, '=O{0}+T{1}/12'.format(row_curr + 2, row_curr + 0))
        set_cell_and_format(ws, 'O{0}'.format(row_curr + 2), border, result.pipe.invert_1)
        set_cell_and_format(ws, 'P{0}'.format(row_curr + 0), border, result.hgl_2)
        set_cell_and_format(ws, 'P{0}'.format(row_curr + 1), border, '=P{0}+T{1}/12'.format(row_curr + 2, row_curr + 0))
        set_cell_and_format(ws, 'P{0}'.format(row_curr + 2), border, result.pipe.invert_2)
        set_cell_and_format(ws, 'Q{0}'.format(row_curr + 0), border, '=sum(O{0}-P{0})'.format(row_curr + 0))
        set_cell_and_format(ws, 'Q{0}'.format(row_curr + 1), border, '=sum(O{0}-P{0})'.format(row_curr + 1))
        set_cell_and_format(ws, 'Q{0}'.format(row_curr + 2), border, '=sum(O{0}-P{0})'.format(row_curr + 2))
        set_cell_and_format(ws, 'R{0}'.format(row_curr + 0), border, result.pipe.section.mannings)
        ws.cell('R{0}'.format(row_curr + 0)).number_format = '0.000'
        set_cell_and_format(ws, 'S{0}'.format(row_curr + 0), border, result.pipe.section.count)
        set_cell_and_format(ws, 'T{0}'.format(row_curr + 0), border, result.pipe.section.rise)
        set_cell_and_format(ws, 'U{0}'.format(row_curr + 0), border, '=Q{0}/B{0}'.format(row_curr + 0))
        ws.cell('U{0}'.format(row_curr + 0)).number_format = '0.00%'
        set_cell_and_format(ws, 'U{0}'.format(row_curr + 2), border, '=Q{0}/B{1}'.format(row_curr + 2, row_curr + 0))
        ws.cell('U{0}'.format(row_curr + 2)).number_format = '0.00%'
        set_cell_and_format(ws, 'V{0}'.format(row_curr + 0), border, result.velocity_actual)
        set_cell_and_format(ws, 'V{0}'.format(row_curr + 2), border, result.velocity_physical)
        set_cell_and_format(ws, 'W{0}'.format(row_curr + 0), border, '=V{0}*PI()*(T{1}/2/12)^2'.format(row_curr + 2,
                                                                                                       row_curr + 0))
        set_cell_and_format(ws, 'X{0}'.format(row_curr + 0), border, '')

    ws.page_setup.fitToWidth = 1
    wb.save(filename)


def set_cell_and_format(ws, coordinate, border, value):
    """Set cell value and format for the given worksheet and coordinate.

    Args:
        ws (openpyxl.worksheet.Worksheet): The worksheet.
        coordinate (str): The coordinate. (e.g. 'B12')
        border (openpyxl.styles.Border): The border size, style, and/or color.
        value: The value to set.

    """
    ws[coordinate] = value
    ws[coordinate].font = Font(name='Arial', size=14)
    ws[coordinate].border = border
    ws[coordinate].alignment = Alignment(horizontal='center', vertical='center')
    ws[coordinate].number_format = '0.00'
